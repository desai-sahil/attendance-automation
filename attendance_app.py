import io
from copy import copy
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import pandas as pd
import streamlit as st


# =========================================================
# App constants
# =========================================================
APP_NAME = "Roll Call"
CREATOR_LINE = "Created by Sahil Desai"
CREATOR_EMAIL = "desai.sahil97@gmail.com"

# Put your logo file here in the repo:
#   assets/big_red_roll_call_logo.png
LOGO_REL_PATH = Path("assets") / "big_red_roll_call_logo.png"


# =========================================================
# Helpers
# =========================================================
def _norm_email(x) -> str:
    return str(x or "").strip().lower()


def _is_blank(x) -> bool:
    return x is None or str(x).strip() == ""


def _cell_text(x) -> str:
    return str(x or "").strip()


def _cell_text_lower(x) -> str:
    return _cell_text(x).lower()


def _find_header_col_ci(ws, header_name: str):
    """Case-insensitive header lookup in row 1. Returns 1-based col index or None."""
    target = str(header_name).strip().lower()
    for cell in ws[1]:
        if cell.value is None:
            continue
        if str(cell.value).strip().lower() == target:
            return cell.column
    return None


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """Copy style/formatting from src_row to dst_row (best-effort)."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def _last_row_with_email(ws, email_col_idx: int, start_row: int = 3) -> int:
    """Find last row (>= start_row) with a non-empty email value."""
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=email_col_idx).value
        if not _is_blank(v):
            last = r
    return last


def _pd_get_ci(row: pd.Series, wanted: str, default=""):
    """Case-insensitive get for pandas Series by column name."""
    wanted_l = wanted.strip().lower()
    for c in row.index:
        if str(c).strip().lower() == wanted_l:
            v = row.get(c, default)
            if pd.isna(v):
                return default
            return v
    return default


def _make_full_and_sortable(first: str, last: str):
    first = str(first or "").strip()
    last = str(last or "").strip()
    full = (first + " " + last).strip()
    if last and first:
        sortable = f"{last}, {first}"
    elif last:
        sortable = last
    else:
        sortable = first
    return full, sortable


def _date_like_equal(a, b: date) -> bool:
    """Compare ws row1 cell (could be datetime/date/string) to a date object b."""
    if a is None:
        return False

    if isinstance(a, datetime):
        return a.date() == b
    if isinstance(a, date):
        return a == b

    s = str(a).strip()
    for fmt in (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m-%d-%Y",
        "%Y-%m-%d",
        "%d-%b",
        "%d-%b-%Y",
        "%b %d, %Y",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == "%d-%b":
                return (dt.month, dt.day) == (b.month, b.day)
            return dt.date() == b
        except Exception:
            pass
    return False


def _set_column_width(ws, col_idx: int, width: float):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _last_used_col_by_headers(ws, header_rows=(1, 2), min_col=1) -> int:
    """
    Returns the last column index that has any non-empty cell in the given header rows.
    Avoids ws.max_column being inflated by formatting.
    """
    last = min_col
    for col in range(min_col, ws.max_column + 1):
        for r in header_rows:
            if not _is_blank(ws.cell(row=r, column=col).value):
                last = col
                break
    return last


def _ensure_lecture_column(ws, lecture_label: str, lecture_date: date) -> int:
    """
    Ensure there is a column whose row2 equals lecture_label (case-insensitive).
    If none exists, append immediately after the last *actually used* header column.

    Enforces formatting:
      - row1: date with "d-mmm" format
      - row2: text
      - rows 3+: integer "0" format (prevents ##### and date rendering)
    """
    lecture_label_clean = lecture_label.strip()
    lecture_label_l = lecture_label_clean.lower()

    # Find existing matches in row 2
    matches = []
    for col in range(1, ws.max_column + 1):
        v2 = ws.cell(row=2, column=col).value
        if _cell_text_lower(v2) == lecture_label_l:
            matches.append(col)

    # Choose best match if multiple
    if len(matches) == 1:
        col_idx = matches[0]
    elif len(matches) > 1:
        col_idx = matches[0]
        for c in matches:
            v1 = ws.cell(row=1, column=c).value
            if _date_like_equal(v1, lecture_date):
                col_idx = c
                break
    else:
        # Append after last REAL header column (not ws.max_column)
        last_used = _last_used_col_by_headers(ws, header_rows=(1, 2), min_col=1)
        col_idx = last_used + 1

    # Write headers
    c1 = ws.cell(row=1, column=col_idx)
    c1.value = lecture_date
    c1.number_format = "d-mmm"  # e.g. 23-Jan

    c2 = ws.cell(row=2, column=col_idx)
    c2.value = lecture_label_clean  # ensures capital L

    _set_column_width(ws, col_idx, 12)

    # Force attendance format for the column
    for r in range(3, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center", vertical="center")

    return col_idx


# =========================================================
# Core logic (GENERALIZED PRESENCE)
# =========================================================
def process_attendance(
    master_file_obj,
    poll_file_obj,
    lecture_number: int,
    lecture_date: date,
    backfill_names_if_blank: bool = True,
):
    """
    GENERALIZED ATTENDANCE LOGIC:
      Present (1) if the student appears in the PollEverywhere export (i.e., their email is listed),
      regardless of whether they answered any question.

    Master sheet expectations:
      Row 1 = date headers for lecture columns (and normal headers for student info cols)
      Row 2 = lecture labels ("Lecture 1", "Lecture 2", ...)
      Row 3+ = student rows
      Student info headers include: Full name, Sortable name, Email (SIS Id optional)
    """
    lecture_label = f"Lecture {int(lecture_number)}"  # capital L enforced

    # --- Read Poll file ---
    try:
        if poll_file_obj.name.lower().endswith(".csv"):
            df_poll = pd.read_csv(poll_file_obj)
        else:
            df_poll = pd.read_excel(poll_file_obj)
    except Exception as e:
        return None, f"Error reading Poll file: {e}"

    # Require Email column in Poll (case-insensitive)
    if not any(str(c).strip().lower() == "email" for c in df_poll.columns):
        return None, "Poll report must contain a column named 'Email'."

    # Build poll roster and presence map
    poll_students = {}   # email -> {"first","last","full","sortable"}
    presence_map = {}    # email -> 1 (present if listed)

    for _, row in df_poll.iterrows():
        email = _norm_email(_pd_get_ci(row, "Email", ""))
        if "@" not in email:
            continue

        first = str(_pd_get_ci(row, "First name", "") or "").strip()
        last = str(_pd_get_ci(row, "Last name", "") or "").strip()
        full, sortable = _make_full_and_sortable(first, last)

        poll_students[email] = {"first": first, "last": last, "full": full, "sortable": sortable}
        presence_map[email] = 1  # listed = present

    if not poll_students:
        return None, "No valid student emails were found in the Poll report."

    # --- Load Master workbook ---
    try:
        wb = openpyxl.load_workbook(master_file_obj)
        ws = wb.active
    except Exception as e:
        return None, f"Error reading Master Excel file: {e}"

    # Roster columns (row 1 headers)
    email_col_idx = _find_header_col_ci(ws, "Email")
    full_name_col_idx = _find_header_col_ci(ws, "Full name")
    sortable_name_col_idx = _find_header_col_ci(ws, "Sortable name")

    if not email_col_idx:
        return None, "Column 'Email' not found in Master Sheet (row 1)."

    # Ensure / create lecture column based on row2 label + row1 date
    lecture_col_idx = _ensure_lecture_column(ws, lecture_label, lecture_date)

    # Map existing master emails -> row
    master_email_to_row = {}
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=email_col_idx).value
        if _is_blank(v):
            continue
        master_email_to_row[_norm_email(v)] = r

    # --- Append NEW students missing from master ---
    last_student_row = _last_row_with_email(ws, email_col_idx, start_row=3)
    append_row = last_student_row + 1
    style_src_row = last_student_row if last_student_row >= 3 else 3
    max_col = ws.max_column

    added_count = 0
    for email, nm in poll_students.items():
        if email in master_email_to_row:
            continue

        if ws.max_row >= style_src_row:
            _copy_row_style(ws, style_src_row, append_row, max_col)

        ws.cell(row=append_row, column=email_col_idx).value = email

        if full_name_col_idx:
            ws.cell(row=append_row, column=full_name_col_idx).value = nm["full"]
        if sortable_name_col_idx:
            ws.cell(row=append_row, column=sortable_name_col_idx).value = nm["sortable"]

        # Initialize attendance for new student
        att_cell = ws.cell(row=append_row, column=lecture_col_idx)
        att_cell.value = 1 if email in presence_map else 0
        att_cell.number_format = "0"
        att_cell.alignment = Alignment(horizontal="center", vertical="center")

        master_email_to_row[email] = append_row
        append_row += 1
        added_count += 1

    # --- Update attendance for everyone in master ---
    updated_present_count = 0
    wrote_zero_count = 0
    backfilled_name_cells = 0

    for email, r in master_email_to_row.items():
        if email in presence_map:
            cell = ws.cell(row=r, column=lecture_col_idx)
            cell.value = 1
            cell.number_format = "0"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            updated_present_count += 1
        else:
            # Only write 0 if blank (preserve manual edits)
            cell = ws.cell(row=r, column=lecture_col_idx)
            if _is_blank(cell.value):
                cell.value = 0
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                wrote_zero_count += 1

        # Optional name backfill (only when blank in master)
        if backfill_names_if_blank and (email in poll_students):
            nm = poll_students[email]
            if full_name_col_idx:
                cur_full = ws.cell(row=r, column=full_name_col_idx).value
                if _is_blank(cur_full) and not _is_blank(nm["full"]):
                    ws.cell(row=r, column=full_name_col_idx).value = nm["full"]
                    backfilled_name_cells += 1
            if sortable_name_col_idx:
                cur_sort = ws.cell(row=r, column=sortable_name_col_idx).value
                if _is_blank(cur_sort) and not _is_blank(nm["sortable"]):
                    ws.cell(row=r, column=sortable_name_col_idx).value = nm["sortable"]
                    backfilled_name_cells += 1

    # --- Save result ---
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    msg = (
        f"Success. Used column '{lecture_label}' dated {lecture_date.strftime('%b %d, %Y')}. "
        f"Present=1 set for {updated_present_count} students (listed in Poll report). "
        f"Absent=0 written for {wrote_zero_count} blank cells. "
        f"Added {added_count} new students. "
    )
    if backfill_names_if_blank:
        msg += f"Backfilled {backfilled_name_cells} name cells."

    return output_buffer, msg


# =========================================================
# Streamlit UI with Apple-inspired Design
# =========================================================

# NOTE: set_page_config must come before any other st.* calls.
st.set_page_config(page_title=APP_NAME, page_icon="📊", layout="centered")

# Custom CSS for Apple-inspired design
st.markdown("""
<style>
    /* Import SF Pro font fallback */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    /* Global styles */
    .stApp {
        background: #f5f5f7;
    }
    
    /* Main content wrapper */
    .main .block-container {
        padding-top: 3rem;
        padding-bottom: 3rem;
        max-width: 980px;
    }
    
    /* Typography overrides */
    h1, h2, h3, h4, h5, h6, p, span, div, label {
        font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Display', 'Inter', system-ui, sans-serif !important;
        letter-spacing: -0.01em;
    }
    
    h1 {
        font-size: 48px !important;
        font-weight: 700 !important;
        color: #1d1d1f !important;
        margin-bottom: 0.5rem !important;
        text-align: center;
    }
    
    /* Card-like containers */
    .stContainer, div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] {
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(20px);
        border-radius: 24px;
        padding: 2rem;
        border: 1px solid rgba(0, 0, 0, 0.06);
        box-shadow: 0 4px 30px rgba(0, 0, 0, 0.05);
    }
    
    /* File uploader styling */
    .stFileUploader {
        background: #fafafa;
        border: 2px dashed #d2d2d7;
        border-radius: 16px;
        padding: 2rem 1.5rem;
        transition: all 0.3s ease;
    }
    
    .stFileUploader:hover {
        border-color: #0071e3;
        background: #f5f5f7;
        transform: translateY(-2px);
    }
    
    .stFileUploader label {
        font-size: 15px !important;
        color: #1d1d1f !important;
        font-weight: 500 !important;
    }
    
    /* Number input styling */
    .stNumberInput input {
        border: 1px solid #d2d2d7;
        border-radius: 12px;
        padding: 14px 16px;
        font-size: 15px;
        transition: all 0.2s;
        background: white;
    }
    
    .stNumberInput input:focus {
        border-color: #0071e3;
        box-shadow: 0 0 0 4px rgba(0, 113, 227, 0.1);
        outline: none;
    }
    
    /* Date input styling */
    .stDateInput input {
        border: 1px solid #d2d2d7;
        border-radius: 12px;
        padding: 14px 16px;
        font-size: 15px;
        transition: all 0.2s;
        background: white;
    }
    
    .stDateInput input:focus {
        border-color: #0071e3;
        box-shadow: 0 0 0 4px rgba(0, 113, 227, 0.1);
        outline: none;
    }
    
    /* Checkbox styling */
    .stCheckbox {
        background: #f5f5f7;
        padding: 1rem;
        border-radius: 12px;
        transition: all 0.2s;
    }
    
    .stCheckbox:hover {
        background: #e8e8ed;
    }
    
    .stCheckbox label {
        font-size: 14px !important;
        color: #1d1d1f !important;
        line-height: 1.5 !important;
    }
    
    /* Button styling */
    .stButton button {
        background: linear-gradient(135deg, #0071e3 0%, #0077ed 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 14px 32px;
        font-size: 17px;
        font-weight: 600;
        transition: all 0.2s;
        width: 100%;
        box-shadow: 0 4px 12px rgba(0, 113, 227, 0.3);
    }
    
    .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0, 113, 227, 0.4);
    }
    
    .stButton button:active {
        transform: translateY(0);
    }
    
    /* Download button */
    .stDownloadButton button {
        background: linear-gradient(135deg, #34c759 0%, #30d158 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 14px 32px;
        font-size: 17px;
        font-weight: 600;
        transition: all 0.2s;
        width: 100%;
        box-shadow: 0 4px 12px rgba(52, 199, 89, 0.3);
    }
    
    .stDownloadButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(52, 199, 89, 0.4);
    }
    
    /* Info box */
    .stInfo {
        background: rgba(0, 113, 227, 0.08);
        border-left: 4px solid #0071e3;
        border-radius: 12px;
        padding: 1rem 1.5rem;
    }
    
    /* Success message */
    .stSuccess {
        background: rgba(52, 199, 89, 0.08);
        border-left: 4px solid #34c759;
        border-radius: 12px;
        padding: 1rem 1.5rem;
    }
    
    /* Error message */
    .stError {
        background: rgba(255, 59, 48, 0.08);
        border-left: 4px solid #ff3b30;
        border-radius: 12px;
        padding: 1rem 1.5rem;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: rgba(255, 255, 255, 0.6);
        border-radius: 12px;
        border: 1px solid rgba(0, 0, 0, 0.05);
        font-size: 15px;
        font-weight: 500;
        color: #1d1d1f;
    }
    
    .streamlit-expanderHeader:hover {
        background: rgba(255, 255, 255, 0.8);
    }
    
    /* Divider */
    hr {
        border: none;
        height: 1px;
        background: rgba(0, 0, 0, 0.08);
        margin: 2rem 0;
    }
    
    /* Subheader styling */
    .stSubheader {
        font-size: 28px !important;
        font-weight: 600 !important;
        color: #1d1d1f !important;
        margin-bottom: 1.5rem !important;
    }
    
    /* Caption/small text */
    .stCaption {
        font-size: 14px !important;
        color: #86868b !important;
    }
    
    /* Section number badges */
    .section-number {
        display: inline-block;
        width: 32px;
        height: 32px;
        background: #f5f5f7;
        border-radius: 8px;
        text-align: center;
        line-height: 32px;
        font-size: 15px;
        font-weight: 600;
        margin-right: 12px;
        color: #1d1d1f;
    }
    
    /* Spinner */
    .stSpinner > div {
        border-color: #0071e3 transparent transparent transparent !important;
    }
</style>
""", unsafe_allow_html=True)

# Header section
st.markdown("""
<div style="text-align: center; margin-bottom: 3rem; animation: fadeIn 0.8s ease-out;">
    <div style="width: 64px; height: 64px; margin: 0 auto 1.5rem; background: linear-gradient(135deg, #FF3B30 0%, #C7001F 100%); border-radius: 18px; display: flex; align-items: center; justify-content: center; box-shadow: 0 8px 30px rgba(255, 59, 48, 0.25);">
        <svg width="36" height="36" viewBox="0 0 24 24" fill="white">
            <path d="M19 3H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zm-7 3c1.93 0 3.5 1.57 3.5 3.5S13.93 13 12 13s-3.5-1.57-3.5-3.5S10.07 6 12 6zm7 13H5v-.23c0-.62.28-1.2.76-1.58C7.47 15.82 9.64 15 12 15s4.53.82 6.24 2.19c.48.38.76.97.76 1.58V19z"/>
        </svg>
    </div>
</div>
""", unsafe_allow_html=True)

st.title(APP_NAME)

st.markdown(f"""
<div style="text-align: center; margin-bottom: 2rem;">
    <p style="font-size: 19px; color: #86868b; margin-bottom: 1rem;">
        Merge attendance reports into a unified master sheet
    </p>
    <p style="font-size: 14px; color: #86868b;">
        {CREATOR_LINE} · <a href="mailto:{CREATOR_EMAIL}" style="color: #0071e3; text-decoration: none;">{CREATOR_EMAIL}</a>
    </p>
</div>
""", unsafe_allow_html=True)

with st.expander("How to use", expanded=False):
    st.markdown("""
    **Confirm emails are included**  
    Both the Master Excel and PollEverywhere export must include a student email column.

    **Download your files**  
    Export the PollEverywhere report as CSV or XLSX, and download your Master Attendance Sheet.

    **Upload & configure**  
    Upload both files, choose the lecture number and date, then process.

    **Attendance logic**  
    A student is marked present (1) if their email appears in the PollEverywhere export. Absences (0) are only written if the cell is blank.
    """)

st.divider()

col1, col2 = st.columns(2)

with col1:
    st.markdown('<h3 style="font-size: 28px; font-weight: 600; margin-bottom: 1.5rem;"><span class="section-number">1</span>Upload Files</h3>', unsafe_allow_html=True)
    master_file = st.file_uploader("Master Excel", type=["xlsx"], key="master")
    poll_file = st.file_uploader("Poll Report", type=["csv", "xlsx"], key="poll")

with col2:
    st.markdown('<h3 style="font-size: 28px; font-weight: 600; margin-bottom: 1.5rem;"><span class="section-number">2</span>Lecture Settings</h3>', unsafe_allow_html=True)
    lecture_number = st.number_input("Lecture number", min_value=1, step=1, value=1)
    lecture_date = st.date_input("Lecture date", value=date.today())

    backfill_names = st.checkbox(
        "Backfill names for existing students (only if blank in master)",
        value=True,
    )

st.divider()

lecture_label_preview = f"Lecture {int(lecture_number)}"

st.info(
    f"**This run will write into:** `{lecture_label_preview}` with date `{lecture_date.strftime('%b %d, %Y')}`"
)

st.markdown("<br>", unsafe_allow_html=True)

if st.button("Process Attendance", type="primary"):
    if not master_file or not poll_file:
        st.error("Please upload both files to continue.")
    else:
        with st.spinner("Processing files..."):
            result_file, message = process_attendance(
                master_file_obj=master_file,
                poll_file_obj=poll_file,
                lecture_number=int(lecture_number),
                lecture_date=lecture_date,
                backfill_names_if_blank=backfill_names,
            )

        if result_file is None:
            st.error(message)
        else:
            st.success(message)

            base = master_file.name.rsplit(".", 1)[0]
            new_filename = f"{base}_UPDATED.xlsx"

            st.download_button(
                label="Download Updated Master Sheet",
                data=result_file,
                file_name=new_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

st.markdown("<br><br>", unsafe_allow_html=True)

st.caption(
    "Keep your Master sheet roster headers consistent (Full name, Sortable name, Email). "
    "The app creates a new lecture column automatically if it doesn't exist."
)
